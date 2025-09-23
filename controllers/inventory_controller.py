# controllers/inventory_controller.py
import openpyxl, os, json
from PyQt6.QtWidgets import QMessageBox, QFileDialog

def get_config_path():
        """Devuelve la ruta del archivo config.json dentro de ProgramData\evoRck"""
        
        program_data = os.environ.get("ProgramData")  # C:\ProgramData
        if not program_data:
            raise EnvironmentError("No se pudo encontrar la carpeta ProgramData")

        evoRck_dir = os.path.join(program_data, "evoRck")
        
        if not os.path.exists(evoRck_dir):
            os.makedirs(evoRck_dir, exist_ok=True)
        config_path = os.path.join(evoRck_dir, "config.json")
        return config_path

def load_inventory_path():
        config_path = get_config_path()
        if os.path.exists(config_path):
            with open(config_path, "r") as f:
                try:
                    data = json.load(f)
                    return data.get("inventory_path", None)
                except json.JSONDecodeError:
                    return None
        return None

def save_inventory_path(file_path):
        config_path = get_config_path()
        data = {"inventory_path": file_path}
        
        with open(config_path, "w") as f:
            json.dump(data, f)

def ask_file_path():
        file_path, _ = QFileDialog.getOpenFileName(
                None,
                "Seleccionar archivo de inventario",
                "",
                "Archivos Excel (*.xlsx *.xls)"
            )
        if not file_path:
                QMessageBox.warning(None, "Cancelado", "No se seleccionó ningún archivo.")
                return {}  # Retornar vacío si canceló
        save_inventory_path(file_path)
        return file_path

class InventoryController:


    def change_file_path():
        ask_file_path()

    def load_inventory(self, sheet_name="Stock"):
        """Carga y agrupa datos del inventario desde Excel, usando persistencia de la ruta."""

        # --- Usar ruta guardada si existe ---
        file_path = load_inventory_path()

        # Si no hay ruta o el archivo no existe, pedir al usuario
        if not file_path or not os.path.exists(file_path):
            file_path,= ask_file_path()
        self.inventory_path = file_path

        # Abrir workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en {file_path}")

        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1]]
        try:
            idx_table = headers.index("TABLE")        # A
            idx_zone = headers.index("ZONE")          # B
            idx_env = headers.index("ENVIRONMENT")    # C
            idx_id = headers.index("ID")              # D
            idx_fur_code = headers.index("FUR CODE")  # F
            idx_stock = headers.index("STOCK")        # H
        except ValueError as e:
            raise ValueError(f"Faltan columnas necesarias: {e}")

        # ---- Extraer todas las filas ----
        all_rows = []
        for row in ws.iter_rows(min_row=2):
            all_rows.append({
                "table": row[idx_table].value,
                "zone": row[idx_zone].value,
                "env": row[idx_env].value,
                "id": row[idx_id].value,
                "fur_code": row[idx_fur_code].value,
                "stock": row[idx_stock].value,
            })

        # ---- Agrupación por FUR CODE ----
        agrupado = {}
        for row in all_rows:
            fur_code = row["fur_code"]
            if not fur_code or fur_code in agrupado:
                continue

            coincidencias = [r for r in all_rows if r["id"] == fur_code]

            zonas = sorted({r["zone"] for r in coincidencias if r["zone"]})
            entornos = sorted({r["env"] for r in coincidencias if r["env"]})
            mesas = sorted({str(r["table"]) for r in coincidencias if r["table"]})

            stock = row["stock"] if isinstance(row["stock"], int) else 0

            agrupado[fur_code] = {
                "stock": stock,
                "zones": zonas,
                "envs": entornos,
                "tables": mesas
            }

        return agrupado, self.inventory_path


    def actualizar_excel(self, path, fur_code: str, cantidad: int, sumar: bool):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Buscar la fila donde está el FUR CODE
        for row in range(2, ws.max_row + 1):
            celda_fur = ws.cell(row=row, column=6)  # Columna F

            if celda_fur.value == fur_code:
                celda_stock = ws.cell(row=row, column=8)  # Columna H
                stock_actual = int(celda_stock.value) if celda_stock.value else 0

                nuevo_stock = stock_actual + cantidad if sumar else stock_actual - cantidad
                nuevo_stock = max(nuevo_stock, 0)  # Evita números negativos
                celda_stock.value = nuevo_stock
                break
        else:
            QMessageBox.warning(self, "Error", f"FUR CODE does not exist: {fur_code}")
            return

        wb.save(path)
        QMessageBox.information(self, "Éxito", f"Stock updated for: {fur_code}.")
