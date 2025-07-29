from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QLineEdit, QPushButton, QHeaderView
)
from PyQt6.QtGui import QPixmap
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QAbstractItemView
from PyQt6.QtWidgets import QMessageBox, QSizePolicy
from views.input_dialog import CustomReasonDialog



import os
import openpyxl


class SecondaryWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Secondary Window")
        self.setGeometry(100, 100, 1200, 600)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        main_layout = QVBoxLayout(main_widget)

        top_layout = QHBoxLayout()

        left_layout = QVBoxLayout()

        self.input1 = QLineEdit()
        self.input1.setPlaceholderText("FUR CODE")
        self.input2 = QLineEdit()
        self.input2.setPlaceholderText("Quantity")

        self.label_info1 = QLabel("Nombre: -")
        self.label_info2 = QLabel("Categoría: -")
        self.label_info3 = QLabel("Ubicación: -")

        self.button1 = QPushButton("Agregar")
        self.button2 = QPushButton("Restar")

        self.button1.clicked.connect(lambda: self.modificar_stock("agregar"))
        self.button2.clicked.connect(lambda: self.modificar_stock("restar"))

        # Contenedor horizontal para los botones
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)

        self.button1.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.button2.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        button_layout.addWidget(self.button1)
        button_layout.addWidget(self.button2)

        # Agregar todos los elementos a la columna izquierda
        for widget in [self.input1, self.input2, self.label_info1, self.label_info2, self.label_info3]:
            left_layout.addWidget(widget)

        left_layout.addLayout(button_layout)  # Agregar el layout de botones al final

        self.excel_path = "src/rcklyt.xlsx"

        self.image_label = QLabel("Table Layout")
        self.image_label.setFixedSize(600, 250)
        self.image_label.setStyleSheet("border: 1px solid gray;")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        top_layout.addLayout(left_layout)
        top_layout.addWidget(self.image_label)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["FUR CODE", "STOCK", "ZONE", "ENVIRONMENT"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Agregar layouts al layout principal
        main_layout.addLayout(top_layout)
        main_layout.addWidget(self.table)

        self.load_inventory_data("src/rcklyt.xlsx")

    def load_inventory_data(self, file_path, sheet_name="Sheet1"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"La hoja '{sheet_name}' no existe.")
            return

        ws = wb[sheet_name]
        self.table.setRowCount(0)

        headers = [cell.value for cell in ws[1]]
        try:
            idx_table = headers.index("TABLE")        # A
            idx_zone = headers.index("ZONE")          # B
            idx_env = headers.index("ENVIRONMENT")    # C
            idx_id = headers.index("ID")              # D
            idx_fur_code = headers.index("FUR CODE")  # F
            idx_stock = headers.index("STOCK")        # H
        except ValueError as e:
            print("Faltan columnas:", e)
            return

        all_rows = []
        for row in ws.iter_rows(min_row=2):
            all_rows.append({
                "table": row[idx_table].value,
                "zone": row[idx_zone].value,
                "env": row[idx_env].value,
                "id": row[idx_id].value,
                "fur_code": row[idx_fur_code].value,
                "stock": row[idx_stock].value,
            })

        agrupado = {}
        for row in all_rows:
            fur_code = row["fur_code"]
            if not fur_code or fur_code in agrupado:
                continue

            coincidencias = [r for r in all_rows if r["id"] == fur_code]

            zonas = sorted({r["zone"] for r in coincidencias if r["zone"]})
            entornos = sorted({r["env"] for r in coincidencias if r["env"]})
            mesas = sorted({str(r["table"]) for r in coincidencias if r["table"]})

            stock = row["stock"] if isinstance(row["stock"], int) else 0

            agrupado[fur_code] = {
                "stock": stock,
                "zones": zonas,
                "envs": entornos,
                "tables": mesas
            }

        self.agrupado_data = agrupado

        for fur_code, data in agrupado.items():
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            self.table.setItem(row_position, 0, QTableWidgetItem(fur_code))
            self.table.setItem(row_position, 1, QTableWidgetItem(str(data["stock"])))
            self.table.setItem(row_position, 2, QTableWidgetItem(" - ".join(data["zones"])))
            self.table.setItem(row_position, 3, QTableWidgetItem(" - ".join(data["envs"])))

        self.table.cellClicked.connect(self.on_cell_clicked)


    def actualizar_stock(self, sumar=True):

        dialogo = CustomReasonDialog()
        if dialogo.exec():  # Si el usuario da OK
            datos = dialogo.get_data()
            print("Datos recibidos:", datos)
            self.label_info1.setText(f"Nombre: {datos['nombre']}")
            self.label_info2.setText(f"Categoría: {datos['categoria']}")
            self.label_info3.setText(f"Ubicación: {datos['ubicacion']}")

            #self.actualizar_stock(sumar=True)
        else:
            print("Cancelado por el usuario.")
        fur_code = self.input1.text().strip()
        cantidad_texto = self.input2.text().strip()

        if not fur_code or not cantidad_texto.isdigit():
            return  # Entradas inválidas

        cantidad = int(cantidad_texto)

        fila_encontrada = -1
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).text() == fur_code:
                fila_encontrada = row
                break

        if fila_encontrada >= 0:
            # Ya existe, actualizar stock
            stock_item = self.table.item(fila_encontrada, 1)
            stock_actual = int(stock_item.text()) if stock_item else 0

            nuevo_stock = stock_actual + cantidad if sumar else stock_actual - cantidad
            nuevo_stock = max(nuevo_stock, 0)  # Evita números negativos

            self.table.setItem(fila_encontrada, 1, QTableWidgetItem(str(nuevo_stock)))
        else:
            # Si no existe, solo permitir agregar
            if sumar:
                nueva_fila = self.table.rowCount()
                self.table.insertRow(nueva_fila)
                self.table.setItem(nueva_fila, 0, QTableWidgetItem(fur_code))
                self.table.setItem(nueva_fila, 1, QTableWidgetItem(str(cantidad)))
                self.table.setItem(nueva_fila, 2, QTableWidgetItem("-"))  # Zona vacía
                self.table.setItem(nueva_fila, 3, QTableWidgetItem("-"))  # Entorno vacío
            else:
                print("No se puede restar a un código inexistente.")
                return

        # Actualizar Excel
        #self.actualizar_excel()
    def on_cell_clicked(self, row, col):
        fur_code_item = self.table.item(row, 0)
        if not fur_code_item:
            return

        fur_code = fur_code_item.text().strip()

        # Autocompletar campo de entrada FUR CODE
        self.input1.setText(fur_code)

        # Enfocar automáticamente el campo de cantidad
        self.input2.setFocus()

        data = self.agrupado_data.get(fur_code, {})

        mesas = data.get("tables", [])
        zonas = data.get("zones", [])
        entornos = data.get("envs", [])

        self.label_info1.setText(f"Nombre: {fur_code}")
        self.label_info2.setWordWrap(True)
        self.label_info2.setText(f"Mesas ({len(mesas)}):\n{', '.join(mesas)}")
        self.label_info2.setMinimumHeight(40)
        self.label_info3.setWordWrap(True)
        self.label_info3.setText(f"Zonas: {', '.join(zonas)} | Entornos: {', '.join(entornos)}")
        self.label_info3.setMinimumHeight(40)

        image_path = os.path.join("src", "table_layouts", f"{fur_code}.jpg")
        if os.path.exists(image_path):
            pixmap = QPixmap(image_path)
            self.image_label.setPixmap(pixmap.scaled(
                self.image_label.size(),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            ))
        else:
            self.image_label.setText("Sin imagen disponible")
            self.image_label.setPixmap(QPixmap())

    def modificar_stock(self, modo: str):
        fur_code = self.input1.text().strip()
        cantidad_texto = self.input2.text().strip()

        if not fur_code or not cantidad_texto.isdigit():
            QMessageBox.warning(self, "Error", "Debe ingresar un FUR CODE y una cantidad válida.")
            return

        cantidad = int(cantidad_texto)

        # Mostrar ventana para seleccionar la razón
        dialogo = CustomReasonDialog(modo=modo, fur_code=fur_code, cantidad=cantidad)
        if dialogo.exec():
            razon = dialogo.get_reason()

            if not razon:
                QMessageBox.warning(self, "Error", "Debe seleccionar una razón.")
                return

            # Mostrar mensaje de confirmación
            mensaje = (
                f"¿Deseas {modo.upper()} \n\nFUR CODE: {fur_code}\nCantidad: {cantidad}\nRazón: {razon}?"
            )

            respuesta = QMessageBox.question(
                self,
                "Confirmar operación",
                mensaje,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if respuesta == QMessageBox.StandardButton.Yes:
                self.razon_actual = razon
                self.actualizar_excel(
                    fur_code=fur_code,
                    cantidad=cantidad,
                    sumar=(modo == "agregar")
                )
        
    def actualizar_excel(self, fur_code: str, cantidad: int, sumar: bool):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        # Buscar la fila donde está el FUR CODE
        for row in range(2, ws.max_row + 1):
            celda_fur = ws.cell(row=row, column=6)  # Columna F

            if celda_fur.value == fur_code:
                celda_stock = ws.cell(row=row, column=8)  # Columna H
                stock_actual = int(celda_stock.value) if celda_stock.value else 0

                nuevo_stock = stock_actual + cantidad if sumar else stock_actual - cantidad
                nuevo_stock = max(nuevo_stock, 0)  # Evita números negativos

                celda_stock.value = nuevo_stock
                break
        else:
            QMessageBox.warning(self, "Error", f"No se encontró el FUR CODE: {fur_code}")
            return

        wb.save(self.excel_path)
        QMessageBox.information(self, "Éxito", f"Stock actualizado correctamente para {fur_code}.")
        self.load_inventory_data(self.excel_path)
